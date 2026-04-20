import openpyxl
import pandas as pd

# Configuration

RAW_FILE = "ETS_ACTUAL_EUTL_data_May2025v2.xlsx"
OUTPUT_FILE = "ets_clean.csv"
 
# Mapping of EU ETS activity type codes to readable sector names
ACTIVITY_MAP = {
    10: "Aviation",
    20: "Combustion of fuels",
    21: "Refining of mineral oil",
    22: "Production of coke",
    23: "Metal ore roasting",
    24: "Production of pig iron or steel",
    25: "Production of cement clinker",
    26: "Production of lime",
    27: "Production of glass",
    28: "Production of ceramic products",
    29: "Production of pulp",
    30: "Production of paper/cardboard",
    31: "Production of carbon black",
    32: "Nitric acid production",
    33: "Adipic acid production",
    34: "Glyoxal/glyoxylic acid production",
    35: "Production of ammonia",
    36: "Production of bulk chemicals",
    37: "Production of hydrogen",
    38: "Production of soda ash",
    39: "Production of petrochemicals",
    40: "Capture of GHGs",
    41: "Transport of GHGs",
    42: "Storage of GHGs",
    43: "Production of aluminium",
    44: "Production of secondary aluminium",
    45: "Production of primary aluminium",
    50: "Shipping",
    99: "Other activities",
}
 
# Mapping of ISO country codes to full country names
COUNTRY_MAP = {
    "AT": "Austria",
    "BE": "Belgium",
    "BG": "Bulgaria",
    "CY": "Cyprus",
    "CZ": "Czech Republic",
    "DE": "Germany",
    "DK": "Denmark",
    "EE": "Estonia",
    "ES": "Spain",
    "FI": "Finland",
    "FR": "France",
    "GB": "United Kingdom",
    "GR": "Greece",
    "HR": "Croatia",
    "HU": "Hungary",
    "IE": "Ireland",
    "IS": "Iceland",
    "IT": "Italy",
    "LI": "Liechtenstein",
    "LT": "Lithuania",
    "LU": "Luxembourg",
    "LV": "Latvia",
    "MT": "Malta",
    "NL": "Netherlands",
    "NO": "Norway",
    "PL": "Poland",
    "PT": "Portugal",
    "RO": "Romania",
    "SE": "Sweden",
    "SI": "Slovenia",
    "SK": "Slovakia",
    "XI": "Northern Ireland",
}
 
# Load raw Excel file
 
print("Loading raw Excel file")
wb = openpyxl.load_workbook(RAW_FILE, read_only=True)
ws = wb.active
 
# Read headers from first row
headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
 
# Select relevant columns
 
print("Selecting relevant columns...")
 
# Only keep EU verified emissions columns (not Swiss CH_ columns)
emission_cols = [h for h in headers if h and str(h).startswith("VERIFIED_EMISSIONS_")]
 
# Columns I am keeping
keep_cols = ["COUNTRY_CODE", "INSTALLATION_NAME", "Final Main Activity Type Code", "CITY"]
all_cols = keep_cols + emission_cols
keep_indices = [headers.index(c) for c in all_cols]
 
# Read all rows, keeping only selected columns
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    rows.append([row[i] for i in keep_indices])
 
df = pd.DataFrame(
    rows,
    columns=["COUNTRY_CODE", "INSTALLATION_NAME", "ACTIVITY_CODE", "CITY"] + emission_cols,
)
 
print(f"   Loaded {len(df):,} installations")
 
# Melt wide format to long format
 
print("Reshaping data from wide to long format...")
 
df_melted = df.melt(
    id_vars=["COUNTRY_CODE", "INSTALLATION_NAME", "ACTIVITY_CODE", "CITY"],
    value_vars=emission_cols,
    var_name="YEAR",
    value_name="VERIFIED_EMISSIONS",
)
 
# Extract year number from column name e.g. "VERIFIED_EMISSIONS_2010" -> 2010
df_melted["YEAR"] = df_melted["YEAR"].str.replace("VERIFIED_EMISSIONS_", "").astype(int)
 
print(f"   Reshaped to {len(df_melted):,} rows")
 
# Remove null and zero emission rows 
 
print("Removing null and zero emission rows...")
 
df_melted = df_melted.dropna(subset=["VERIFIED_EMISSIONS"])
df_melted = df_melted[df_melted["VERIFIED_EMISSIONS"] > 0]
 
print(f"   {len(df_melted):,} rows remaining after cleaning")
 
# Map activity codes to sector names 
 
print("Mapping activity codes to sector names...")
 
df_melted["SECTOR"] = (
    df_melted["ACTIVITY_CODE"]
    .map(ACTIVITY_MAP)
    .fillna("Other activities")
)
 
# Map country codes to full country names
 
print("Mapping country codes to full country names...")
 
df_melted["COUNTRY"] = (
    df_melted["COUNTRY_CODE"]
    .map(COUNTRY_MAP)
    .fillna(df_melted["COUNTRY_CODE"])
)
 
# Select final columns and save 
 
print("Saving cleaned dataset...")
 
df_clean = df_melted[
    [
        "COUNTRY_CODE",
        "COUNTRY",
        "INSTALLATION_NAME",
        "CITY",
        "SECTOR",
        "YEAR",
        "VERIFIED_EMISSIONS",
    ]
].copy()
 
df_clean = df_clean.sort_values(["COUNTRY", "YEAR"]).reset_index(drop=True)
 
df_clean.to_csv(OUTPUT_FILE, index=False)
 
# Summary

print("\n Data cleaning complete!")
print(f"Output file  : {OUTPUT_FILE}")
print(f"Total rows   : {len(df_clean):,}")
print(f"Years        : {df_clean['YEAR'].min()} to {df_clean['YEAR'].max()}")
print(f"Countries    : {df_clean['COUNTRY'].nunique()}")
print(f"Sectors      : {df_clean['SECTOR'].nunique()}")
print(f"Installations: {df_clean['INSTALLATION_NAME'].nunique():,}")